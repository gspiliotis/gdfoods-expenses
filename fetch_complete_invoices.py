#!/usr/bin/env python3
"""
Fetch all invoices from myDATA API and write detailed line items to Excel.
Excludes credit note invoice types (5, 5.1, 5.2).
"""
import argparse
import os
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from typing import List, Dict, Optional, Tuple
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Load environment variables from .env file
load_dotenv()

# API Constants
USER_ID = os.getenv("MYDATA_USER_ID")
API_KEY = os.getenv("MYDATA_API_KEY")
API_BASE_URL = "https://mydatapi.aade.gr/myDATA/RequestDocs"

MEASUREMENT_UNIT_MAP = {
    "1": "Τεμάχια",
    "2": "Κιλά",
    "3": "Λίτρα",
    "4": "Μέτρα",
    "5": "Τετραγωνικά Μέτρα",
    "6": "Κυβικά Μέτρα",
    "7": "Λοιπές Περιπτώσεις",
}

VAT_CATEGORY_MAP = {
    "1": "24%",
    "2": "13%",
    "3": "6%",
    "4": "17%",
    "5": "9%",
    "6": "4%",
    "7": "0%",
    "8": "-",
    "9": "3%",
    "10": "4%",
}

if not USER_ID or not API_KEY:
    print("Error: MYDATA_USER_ID and MYDATA_API_KEY environment variables must be set", file=sys.stderr)
    sys.exit(1)


def convert_date_to_api_format(date_str: str) -> str:
    """Convert date from YYYY-MM-DD to DD/MM/YYYY format for API."""
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    return date_obj.strftime("%d/%m/%Y")


def fetch_invoices(
    date_from: str,
    date_to: str,
    next_partition_key: Optional[str] = None,
    next_row_key: Optional[str] = None
) -> str:
    """Fetch invoices from myDATA API."""
    api_date_from = convert_date_to_api_format(date_from)
    api_date_to = convert_date_to_api_format(date_to)

    params = {
        "mark": "1",
        "dateFrom": api_date_from,
        "dateTo": api_date_to
    }

    if next_partition_key:
        params["nextPartitionKey"] = next_partition_key
    if next_row_key:
        params["nextRowKey"] = next_row_key

    headers = {
        "aade-user-id": USER_ID,
        "Ocp-Apim-Subscription-Key": API_KEY
    }

    try:
        response = requests.get(API_BASE_URL, params=params, headers=headers, timeout=30)
        response.raise_for_status()
        return response.text
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data: {e}", file=sys.stderr)
        if hasattr(e, 'response') and e.response is not None:
            print(f"Response text: {e.response.text}", file=sys.stderr)
        return ""


def parse_invoices(xml_content: str) -> Tuple[List[Dict], Optional[str], Optional[str]]:
    """
    Parse XML response and extract invoice line item rows.
    Skips invoices with invoice_type in ["5", "5.1", "5.2"].
    Each invoiceDetails element produces one row with the invoice-level fields repeated.

    Returns:
        Tuple of (list of row dicts, next_partition_key, next_row_key)
    """
    if not xml_content:
        return [], None, None

    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError as e:
        print(f"Error parsing XML: {e}", file=sys.stderr)
        return [], None, None

    ns = {'ns': 'http://www.aade.gr/myDATA/invoice/v1.0'}

    # Extract pagination tokens
    next_partition_key = None
    next_row_key = None

    continuation_token = root.find("ns:continuationToken", ns)
    if continuation_token is not None:
        npk = continuation_token.find("ns:nextPartitionKey", ns)
        nrk = continuation_token.find("ns:nextRowKey", ns)
        if npk is not None:
            next_partition_key = npk.text
        if nrk is not None:
            next_row_key = nrk.text

    rows = []
    excluded_types = {"5", "5.1", "5.2"}

    invoices_doc = root.find("ns:invoicesDoc", ns)
    if invoices_doc is None:
        return rows, next_partition_key, next_row_key

    for invoice in invoices_doc.findall("ns:invoice", ns):
        # Get invoice header and check type
        invoice_header = invoice.find("ns:invoiceHeader", ns)
        if invoice_header is None:
            continue

        invoice_type_elem = invoice_header.find("ns:invoiceType", ns)
        invoice_type = invoice_type_elem.text.strip() if invoice_type_elem is not None and invoice_type_elem.text else ""

        if invoice_type in excluded_types:
            continue

        # Get issuer information
        issuer = invoice.find("ns:issuer", ns)
        issuer_vat = ""
        issuer_name = ""
        if issuer is not None:
            vat_elem = issuer.find("ns:vatNumber", ns)
            name_elem = issuer.find("ns:name", ns)
            issuer_vat = vat_elem.text.strip() if vat_elem is not None and vat_elem.text else ""
            issuer_name = name_elem.text.strip() if name_elem is not None and name_elem.text else ""

        # Get header fields
        issue_date_elem = invoice_header.find("ns:issueDate", ns)
        series_elem = invoice_header.find("ns:series", ns)
        aa_elem = invoice_header.find("ns:aa", ns)

        issue_date = issue_date_elem.text.strip() if issue_date_elem is not None and issue_date_elem.text else ""
        series = series_elem.text.strip() if series_elem is not None and series_elem.text else ""
        aa = aa_elem.text.strip() if aa_elem is not None and aa_elem.text else ""

        # Get all invoiceDetails elements
        invoice_details_list = invoice.findall("ns:invoiceDetails", ns)

        if not invoice_details_list:
            # Invoice with no details: emit one row with blank detail fields
            rows.append({
                "issuer_vat": issuer_vat,
                "issuer_name": issuer_name,
                "issue_date": issue_date,
                "series": series,
                "aa": aa,
                "item_code": "",
                "item_descr": "",
                "quantity": None,
                "measurement_unit": "",
                "net_value": None,
                "vat_category": "",
                "vat_amount": None,
            })
            continue

        for detail in invoice_details_list:
            item_code_elem = detail.find("ns:itemCode", ns)
            item_descr_elem = detail.find("ns:itemDescr", ns)
            quantity_elem = detail.find("ns:quantity", ns)
            measurement_unit_elem = detail.find("ns:measurementUnit", ns)
            net_value_elem = detail.find("ns:netValue", ns)
            vat_category_elem = detail.find("ns:vatCategory", ns)
            vat_amount_elem = detail.find("ns:vatAmount", ns)

            # Parse numeric fields
            quantity = None
            if quantity_elem is not None and quantity_elem.text:
                try:
                    quantity = float(quantity_elem.text.strip())
                except ValueError:
                    pass

            net_value = None
            if net_value_elem is not None and net_value_elem.text:
                try:
                    net_value = float(net_value_elem.text.strip())
                except ValueError:
                    pass

            vat_amount = None
            if vat_amount_elem is not None and vat_amount_elem.text:
                try:
                    vat_amount = float(vat_amount_elem.text.strip())
                except ValueError:
                    pass

            # Map measurement unit code to text
            mu_raw = measurement_unit_elem.text.strip() if measurement_unit_elem is not None and measurement_unit_elem.text else ""
            measurement_unit = MEASUREMENT_UNIT_MAP.get(mu_raw, "Άγνωστο") if mu_raw else ""

            # Map VAT category code to text
            vc_raw = vat_category_elem.text.strip() if vat_category_elem is not None and vat_category_elem.text else ""
            vat_category = VAT_CATEGORY_MAP.get(vc_raw, vc_raw)

            rows.append({
                "issuer_vat": issuer_vat,
                "issuer_name": issuer_name,
                "issue_date": issue_date,
                "series": series,
                "aa": aa,
                "item_code": item_code_elem.text.strip() if item_code_elem is not None and item_code_elem.text else "",
                "item_descr": item_descr_elem.text.strip() if item_descr_elem is not None and item_descr_elem.text else "",
                "quantity": quantity,
                "measurement_unit": measurement_unit,
                "net_value": net_value,
                "vat_category": vat_category,
                "vat_amount": vat_amount,
            })

    return rows, next_partition_key, next_row_key


def fetch_all_invoices(date_from: str, date_to: str) -> List[Dict]:
    """Fetch all invoices with pagination."""
    print(f"Fetching all invoices for date range: {date_from} to {date_to}")

    all_rows = []
    next_partition_key = None
    next_row_key = None
    page = 1

    while True:
        xml_content = fetch_invoices(
            date_from, date_to,
            next_partition_key=next_partition_key,
            next_row_key=next_row_key
        )

        if not xml_content:
            break

        rows, next_partition_key, next_row_key = parse_invoices(xml_content)
        all_rows.extend(rows)

        print(f"  Page {page}: {len(rows)} row(s)")
        page += 1

        if not next_partition_key or not next_row_key:
            break

    print(f"Total rows: {len(all_rows)}")
    return all_rows


def write_excel(rows: List[Dict], output_file: str):
    """Write rows to Excel file with proper formatting."""
    fieldnames = [
        "issuer_vat", "issuer_name", "issue_date", "series", "aa",
        "item_code", "item_descr", "quantity", "measurement_unit",
        "net_value", "vat_category", "vat_amount"
    ]
    header = [
        "Issuer VAT", "Issuer Name", "Issue Date", "Series", "AA",
        "Item Code", "Item Description", "Quantity", "Measurement Unit",
        "Net Value", "VAT Category", "VAT Amount"
    ]
    # Columns that should be numeric (0-indexed within fieldnames)
    numeric_fields = {"quantity", "net_value", "vat_amount"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # Write header
    for col_idx, col_name in enumerate(header, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Format all columns as text by default
    for col_idx in range(1, len(fieldnames) + 1):
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            cell.number_format = '@'

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            value = row[field]
            cell = ws.cell(row=row_idx, column=col_idx)
            if field in numeric_fields:
                # Write as number (None becomes empty cell)
                if value is not None:
                    cell.value = value
                    cell.number_format = '0.00'
                # else leave cell empty
            else:
                cell.value = str(value) if value else ""
                cell.number_format = '@'

    wb.save(output_file)
    print(f"Wrote {len(rows)} row(s) to {output_file}")


def validate_date(date_str: str) -> bool:
    """Validate date format (YYYY-MM-DD)."""
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def main():
    parser = argparse.ArgumentParser(
        description="Fetch all invoices from myDATA API and write detailed line items to Excel"
    )
    parser.add_argument(
        "--from-date",
        help="Start date in YYYY-MM-DD format (default: today)"
    )
    parser.add_argument(
        "--to-date",
        help="End date in YYYY-MM-DD format (default: today)"
    )
    parser.add_argument(
        "--output",
        default="complete_invoices.xlsx",
        help="Output Excel file (default: complete_invoices.xlsx)"
    )

    args = parser.parse_args()

    today = datetime.now().strftime("%Y-%m-%d")
    date_from = args.from_date if args.from_date else today
    date_to = args.to_date if args.to_date else today

    if not validate_date(date_from):
        print(f"Error: Invalid start date '{date_from}'. Use YYYY-MM-DD format.", file=sys.stderr)
        sys.exit(1)

    if not validate_date(date_to):
        print(f"Error: Invalid end date '{date_to}'. Use YYYY-MM-DD format.", file=sys.stderr)
        sys.exit(1)

    rows = fetch_all_invoices(date_from, date_to)

    if not rows:
        print("\nNo invoice data found")
        sys.exit(0)

    write_excel(rows, args.output)
    print("\nDone!")


if __name__ == "__main__":
    main()
